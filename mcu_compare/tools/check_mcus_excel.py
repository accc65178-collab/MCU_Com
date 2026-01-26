import os
import sys
import json
from glob import glob

try:
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
except Exception as e:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

# Inputs
EXCEL_PATH = r"E:\MCU-com\mcus.xlsx"
DATA_DIR = r"E:\MCU-com\data"
SHEET_NAME = None  # None = active sheet
COLUMN = "A"  # Column with part numbers


def read_excel_part_numbers(path: str, sheet_name=None, column="A"):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    parts = []
    # Determine column index (1-based)
    if isinstance(column, int):
        col_idx = max(1, int(column))
    else:
        col_idx = column_index_from_string(str(column))
    # Use iter_rows for ReadOnlyWorksheet
    for row in ws.iter_rows(min_row=1, min_col=col_idx, max_col=col_idx, values_only=True):
        val = row[0]
        if val is None:
            continue
        s = str(val).strip()
        if s:
            parts.append(s)
    return parts


def load_all_mcu_names(data_dir: str):
    names = []
    for fp in glob(os.path.join(data_dir, "mcus_*.json")):
        try:
            with open(fp, "r", encoding="utf-8") as f:
                arr = json.load(f)
            if isinstance(arr, dict):
                # legacy structure not expected, skip
                continue
            for rec in arr:
                name = str(rec.get("name", "")).strip()
                if name:
                    names.append(name)
        except Exception:
            continue
    return names


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found: {EXCEL_PATH}", file=sys.stderr)
        sys.exit(1)
    excel_names = read_excel_part_numbers(EXCEL_PATH, SHEET_NAME, COLUMN)
    data_names = load_all_mcu_names(DATA_DIR)

    # Normalization: lowercase and remove non-alphanumeric characters
    import re
    def norm(s: str) -> str:
        return re.sub(r"[^a-z0-9]", "", (s or "").strip().lower())

    data_set = {norm(n): n for n in data_names}

    missing = []
    for n in excel_names:
        key = norm(n)
        if key not in data_set:
            missing.append(n)

    print(f"Total in Excel: {len(excel_names)}")
    print(f"Total in data:  {len(data_names)}")
    print(f"Missing count:  {len(missing)}")
    if missing:
        print("\nMissing part numbers:")
        for m in missing:
            print(m)

    # Write report
    out_path = os.path.join(os.path.dirname(EXCEL_PATH), "missing_from_excel.txt")
    try:
        with open(out_path, "w", encoding="utf-8") as f:
            f.write("\n".join(missing))
        print(f"\nReport written: {out_path}")
    except Exception as e:
        print(f"WARNING: could not write report: {e}")


if __name__ == "__main__":
    main()
